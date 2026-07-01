#!/usr/bin/env python3
# reviewer.py - purple-team review workbook across out/<id>/verdict.json (+ new.json)
#
# Usage (from repo root, using the project venv):
#   venv/Scripts/python.exe auto-scripts/reviewer.py
#   venv/Scripts/python.exe auto-scripts/reviewer.py --out-dir out --xlsx review.xlsx
#
# Same classification as stats-verdicts.py:
#   covered   -> root_causes is not null (existing CRS root-cause rule already fires)
#   gated     -> scope_gate.decision in HALT_GATES (virtual-patch-only / out-of-scope-structural)
#   in-scope  -> proceeded to Step 3 (crs-rule-author); new.json present if authored
#   <other>   -> unexpected scope_gate.decision value, surfaced as-is
#
# For every CVE this also derives a purple-team recommendation line: what a
# reviewer should actually do with the case (accept CRS baseline, ship the new
# rule, virtual-patch outside CRS core, or push back to app/vendor), grounded
# in the skill's own rationale/note/recommendation text so nothing is invented.
#
# Output: one .xlsx workbook with two sheets - "Summary" (tallies) and
# "CVE Review" (one row per CVE with the recommendation).
#
# For handoff to the content/purple team, staging + zipping the review
# artifacts per CVE lives in resources-compress.py (split out to keep this
# script focused on the xlsx report):
#   venv/Scripts/python.exe auto-scripts/resources-compress.py
#
# resources-compress.py zips each CVE's artifacts under a "<cve_id>/<filename>"
# path (e.g. "CVE-2025-0108/verdict.json"), regardless of what the zip itself
# is named or where the recipient extracts it. The "verdict_ref"/"new_rule_ref"
# columns below use that same relative layout, so a reviewer with the handoff
# zip open in one window and this xlsx in another can jump straight to the
# backing file without needing to know the zip's name or extraction path.

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Must match run-pipeline.py / stats-verdicts.py.
HALT_GATES = {"virtual-patch-only", "out-of-scope-structural"}

parser = argparse.ArgumentParser(description="Purple-team review workbook from CRS verdict.json/new.json")
parser.add_argument("--out-dir", default="out", help="Directory holding <id>/verdict.json (default: out)")
parser.add_argument("--xlsx", default="cve-review.xlsx", help="Output .xlsx path (default: cve-review.xlsx)")
args = parser.parse_args()

out_dir = Path(args.out_dir)
verdict_paths = sorted(out_dir.glob("*/verdict.json"))


def load_json(p: Path):
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def rule_list(rules, key="id"):
    return ", ".join(str(r.get(key)) for r in rules if r.get(key) is not None)


def classify(v: dict):
    """Return (status, gate_decision) - same rule as stats-verdicts.py."""
    decision = (v.get("scope_gate") or {}).get("decision")
    if v.get("root_causes") is not None:
        return "covered", None
    if decision in HALT_GATES:
        return "gated", decision
    if decision == "in-scope":
        return "in-scope", decision
    if decision is None:
        return "no-gate", None
    return decision, decision


def recommendation_for(status: str, gate_decision, v: dict, new: dict | None):
    """Purple-team recommendation text, grounded in the skill's own output."""
    if status == "covered":
        rc = v.get("root_causes") or {}
        rules = rc.get("root_cause_rules") or []
        ids = rule_list(rules)
        rec = rc.get("recommendation") or {}
        pl_coverage = rec.get("pl_coverage", "")
        return (
            f"ACCEPT baseline - CRS root-cause rule(s) {ids or 'n/a'} already block this exploit. "
            f"No new rule needed; keep current PL. {pl_coverage}"
        ).strip()

    if status == "in-scope":
        if new and new.get("rule"):
            r = new["rule"]
            rid = r.get("id", "n/a")
            msg = None
            secrule = r.get("secrule_text", "")
            if "msg:'" in secrule:
                msg = secrule.split("msg:'", 1)[1].split("'", 1)[0]
            target = new.get("target_file", "n/a")
            pl = r.get("paranoia_level", "?")
            sev = r.get("severity", "?")
            conf = new.get("confidence", "?")
            return (
                f"REVIEW & SHIP candidate rule {rid} ({msg or 'see new.json'}) at PL{pl}/{sev} "
                f"for {target}. Author confidence: {conf}. Validate against FP corpus before "
                f"merging into coreruleset/rules/{target}."
            )
        return (
            "IN-SCOPE but Step 3 (crs-rule-author) has not produced new.json yet - "
            "rerun crs-rule-author for this template before closing out."
        )

    if gate_decision == "virtual-patch-only":
        note = v.get("note") or (v.get("scope_gate") or {}).get("rationale", "")
        return (
            "OUT OF CRS-CORE SCOPE - deterministic virtual patch is possible but is "
            f"app/device-specific, not generic CRS material. {note} "
            "Recommend: ship as a custom/virtual-patch rule in a site-specific ruleset "
            "(not coreruleset), or push the app owner/vendor for an upstream fix."
        )

    if gate_decision == "out-of-scope-structural":
        note = v.get("note") or (v.get("scope_gate") or {}).get("rationale", "")
        return (
            "OUT OF WAF SCOPE (business logic) - no content signature distinguishes this "
            f"from a legitimate request. {note} "
            "Recommend: fix at the application layer (authz/token/crypto logic review, "
            "vendor patch) - a WAF signature cannot reliably catch this class."
        )

    return f"MANUAL REVIEW - unexpected scope_gate.decision={gate_decision!r}; inspect verdict.json directly."


def rule_logic(status: str, v: dict, new: dict | None):
    """(short structural summary, prose explanation) - never the raw pattern/secrule_text.

    covered  -> root_causes.recommendation.rule_analysis[] (operator/transforms/matched_at,
                pattern_excerpt + trigger_explanation)
    in-scope -> new.json rule.{operator,variables,transforms} + design_rationale.operator
    """
    if status == "covered":
        analyses = ((v.get("root_causes") or {}).get("recommendation") or {}).get("rule_analysis") or []
        short_parts, expl_parts = [], []
        for a in analyses:
            rid = a.get("id", "?")
            op = a.get("operator", "")
            matched_at = a.get("matched_at", "")
            transforms = ",".join(a.get("transforms", []) or [])
            short_parts.append(f"{rid}: {op} on {matched_at} [{transforms}]")
            excerpt = a.get("pattern_excerpt", "")
            trig = a.get("trigger_explanation", "")
            expl_parts.append(f"{rid}: {excerpt} {trig}".strip())
        return "; ".join(short_parts), "\n".join(expl_parts)

    if status == "in-scope" and new and new.get("rule"):
        r = new["rule"]
        op = r.get("operator", "")
        varz = r.get("variables", "")
        transforms = ",".join(r.get("transforms", []) or [])
        short = f"{op} on {varz} [{transforms}]"
        expl = (new.get("design_rationale") or {}).get("operator", "")
        return short, expl

    return "", ""


rows = []
status_count = Counter()
gate_count = Counter()
family_count = Counter()
blocked_count = Counter()
authored_count = Counter()

for vp in verdict_paths:
    cid = vp.parent.name
    v = load_json(vp)
    if v is None:
        rows.append({
            "cve_id": cid, "status": "parse-error", "gate_decision": "",
            "families": "", "template_path": "", "injection_point": "",
            "protocol": "", "probe_blocked": "", "anomaly_inbound": "",
            "anomaly_threshold": "", "root_cause_rules": "", "new_rule_id": "",
            "new_rule_msg": "", "new_rule_pl": "", "new_rule_severity": "",
            "new_rule_target_file": "", "rule_logic": "", "rule_logic_explanation": "",
            "recommendation": "verdict.json failed to parse - inspect manually.",
            "evidence": "", "source_verdict": str(vp),
        })
        status_count["parse-error"] += 1
        continue

    status, gate_decision = classify(v)
    status_count[status] += 1
    if gate_decision in HALT_GATES:
        gate_count[gate_decision] += 1

    classification = v.get("classification") or {}
    for fam in classification.get("families", []) or []:
        family_count[fam] += 1

    probe = v.get("probe") or {}
    score = probe.get("anomaly_score") or {}
    if "blocked" in probe:
        blocked_count["blocked" if probe.get("blocked") else "not-blocked"] += 1

    new = None
    new_path = vp.parent / "new.json"
    if status == "in-scope":
        new = load_json(new_path)
        authored_count["authored" if new else "pending"] += 1

    root_cause_rules = ((v.get("root_causes") or {}).get("root_cause_rules")) or []

    new_rule = (new or {}).get("rule") or {}
    new_msg = ""
    secrule = new_rule.get("secrule_text", "")
    if "msg:'" in secrule:
        new_msg = secrule.split("msg:'", 1)[1].split("'", 1)[0]

    evidence = (v.get("scope_gate") or {}).get("rationale") or v.get("note") or ""
    logic_short, logic_expl = rule_logic(status, v, new)

    # Relative to the handoff zip root (resources-compress.py layout), not the
    # local out/ tree - stays valid however the zip is named or extracted.
    verdict_ref = f"{cid}/verdict.json"
    new_rule_ref = f"{cid}/new.json" if new else ""

    rows.append({
        "cve_id": cid,
        "status": status,
        "gate_decision": gate_decision or "",
        "families": ", ".join(classification.get("families", []) or []),
        "template_path": v.get("template_path", ""),
        "injection_point": classification.get("injection_point", ""),
        "protocol": classification.get("protocol", ""),
        "probe_blocked": probe.get("blocked", ""),
        "anomaly_inbound": score.get("inbound", ""),
        "anomaly_threshold": score.get("threshold", ""),
        "root_cause_rules": rule_list(root_cause_rules),
        "new_rule_id": new_rule.get("id", ""),
        "new_rule_msg": new_msg,
        "new_rule_pl": new_rule.get("paranoia_level", ""),
        "new_rule_severity": new_rule.get("severity", ""),
        "new_rule_target_file": new.get("target_file", "") if new else "",
        "rule_logic": logic_short,
        "rule_logic_explanation": logic_expl,
        "recommendation": recommendation_for(status, gate_decision, v, new),
        "evidence": evidence,
        "verdict_ref": verdict_ref,
        "new_rule_ref": new_rule_ref,
        "source_verdict": str(vp),
    })

total = len(rows)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

STATUS_FILL = {
    "covered":     PatternFill("solid", fgColor="C6EFCE"),
    "in-scope":    PatternFill("solid", fgColor="FFEB9C"),
    "gated":       PatternFill("solid", fgColor="FFC7CE"),
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


# --- Sheet 1: Summary ---
ws = wb.active
ws.title = "Summary"
ws.append(["Metric", "Value"])
style_header(ws, 2)
ws.append(["Total CVEs reviewed", total])
ws.append([])
ws.append(["Status", "Count", "Pct"])
for st, n in status_count.most_common():
    ws.append([st, n, f"{n/total*100:.1f}%" if total else ""])
ws.append([])
ws.append(["Gate breakdown (skipped Step 2 & 3)", "Count"])
for g, n in gate_count.most_common():
    ws.append([g, n])
ws.append([])
ws.append(["In-scope outcome (Step 3 new.json)", "Count"])
for a, n in authored_count.most_common():
    ws.append([a, n])
ws.append([])
ws.append(["Probe blocked at PL2", "Count"])
for b, n in blocked_count.most_common():
    ws.append([b, n])
ws.append([])
ws.append(["Attack family", "Count"])
for fam, n in family_count.most_common():
    ws.append([fam, n])

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 10

# --- Sheet 2: CVE Review ---
ws2 = wb.create_sheet("CVE Review")
columns = [
    ("cve_id",               "CVE ID",                14),
    ("status",               "Status",                12),
    ("gate_decision",        "Gate decision",          22),
    ("families",             "Attack families",        24),
    ("template_path",        "Nuclei template",        42),
    ("injection_point",      "Injection point",        45),
    ("protocol",             "Protocol",               10),
    ("probe_blocked",        "Blocked @PL2 probe",     14),
    ("anomaly_inbound",      "Inbound score",          12),
    ("anomaly_threshold",    "Threshold",              10),
    ("root_cause_rules",     "Root-cause rule id(s)",  16),
    ("new_rule_id",          "New rule id",            12),
    ("new_rule_msg",         "New rule msg",           30),
    ("new_rule_pl",          "New rule PL",            10),
    ("new_rule_severity",    "New rule severity",      12),
    ("new_rule_target_file", "New rule target file",   32),
    ("rule_logic",           "Rule logic (operator/vars/transforms)", 40),
    ("rule_logic_explanation", "Rule logic explanation", 60),
    ("recommendation",       "Purple team recommendation", 70),
    ("evidence",             "Evidence (rationale/note)",  70),
    ("verdict_ref",          "Handoff ref: verdict.json", 28),
    ("new_rule_ref",         "Handoff ref: new.json",  24),
    ("source_verdict",       "Source verdict.json",   30),
]
ws2.append([label for _, label, _ in columns])
style_header(ws2, len(columns))

for r in rows:
    ws2.append([r.get(key, "") for key, _, _ in columns])

status_col = 2  # "status" is column B
for i, r in enumerate(rows, start=2):
    fill = STATUS_FILL.get(r["status"])
    if fill:
        ws2.cell(row=i, column=status_col).fill = fill
    for c in range(1, len(columns) + 1):
        ws2.cell(row=i, column=c).alignment = WRAP

for idx, (_, _, width) in enumerate(columns, start=1):
    ws2.column_dimensions[get_column_letter(idx)].width = width

ws2.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

wb.save(args.xlsx)
print(f"wrote {total} CVE row(s) -> {args.xlsx}")
print(f"  status: {dict(status_count)}")
print(f"  gate breakdown: {dict(gate_count)}")
print(f"  in-scope outcome: {dict(authored_count)}")
